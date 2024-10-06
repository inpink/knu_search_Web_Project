import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

def saveQueryValue(query_rows,query_col_start,query_data,sheet_data):
  # 각 쿼리의 P@1 값을 2차원 리스트에 저장
  for i, row in enumerate(query_rows):
    query_data[i][0].append(sheet_data.iloc[row - 1, query_col_start])  # KNU
    query_data[i][1].append(sheet_data.iloc[row - 1, query_col_start+1])  # ResNetTrans
    query_data[i][2].append(sheet_data.iloc[row - 1, query_col_start+2])  # BM25
    query_data[i][3].append(sheet_data.iloc[row - 1, query_col_start+3])  # BM25+ResNetTrans

def printPK(name, columns, query_averages_p1):
  print(f"{name} Averages:")
  for i in range(10):
    print(f"Query {i+1}:")
    for j in range(4):
      print(f"  {columns[j]}: {query_averages_p1[i][j]}")
  print()

def solution():
  # 파일 경로 지정
  file_path = '질문지_10.xlsx'

  # 각 쿼리별로 P@1 ~ P@5 값들이 위치한 셀 정보 (P@1: 10에서 시작, P@2: 11, P@3: 12, P@4: 13, P@5: 14)
  query_rows_p1 = [10 + i * 12 for i in range(10)]  # P@1용 행 번호 리스트
  query_rows_p2 = [11 + i * 12 for i in range(10)]  # P@2용 행 번호 리스트
  query_rows_p3 = [12 + i * 12 for i in range(10)]  # P@3용 행 번호 리스트
  query_rows_p4 = [13 + i * 12 for i in range(10)]  # P@4용 행 번호 리스트
  query_rows_p5 = [14 + i * 12 for i in range(10)]  # P@5용 행 번호 리스트
  query_rows_ap5 = [15 + i * 12 for i in range(10)]  # AP@5용 행 번호 리스트
  query_rows_ndcg = [5 + i * 12 for i in range(10)]  # ndcg용 행 번호 리스트

  query_col_start_p1 = 7  # P@1 시작 열 번호
  query_col_start_p2 = 7  # P@2 시작 열 번호
  query_col_start_p3 = 7  # P@3 시작 열 번호
  query_col_start_p4 = 7  # P@4 시작 열 번호
  query_col_start_p5 = 7  # P@5 시작 열 번호
  query_col_start_ap5 = 7  # AP@5 시작 열 번호
  query_col_start_ndcg = 17  # ndcg 시작

  # 쿼리별로 각 항목의 P@1 ~ P@5 값을 저장할 2차원 리스트 (총 10개의 쿼리, 4개의 항목)
  query_data_p1 = [[[] for _ in range(4)] for _ in range(10)]  # P@1용 10 x 4 배열
  query_data_p2 = [[[] for _ in range(4)] for _ in range(10)]  # P@2용 10 x 4 배열
  query_data_p3 = [[[] for _ in range(4)] for _ in range(10)]  # P@3용 10 x 4 배열
  query_data_p4 = [[[] for _ in range(4)] for _ in range(10)]  # P@4용 10 x 4 배열
  query_data_p5 = [[[] for _ in range(4)] for _ in range(10)]  # P@5용 10 x 4 배열
  query_data_ap5 = [[[] for _ in range(4)] for _ in range(10)]  # P@5용 10 x 4 배열
  query_data_ndcg = [[[] for _ in range(4)] for _ in range(10)]  # ndcg용 10 x 4 배열


  # Sheet1부터 Sheet10까지 읽기
  for sheet_num in range(1, 11):
    sheet_name = f'Sheet{sheet_num}'
    # 해당 시트 읽기
    sheet_data = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

    # 각 쿼리의 P@1~5 값을 2차원 리스트에 저장
    saveQueryValue(query_rows_p1, query_col_start_p1, query_data_p1, sheet_data)
    saveQueryValue(query_rows_p2, query_col_start_p2, query_data_p2, sheet_data)
    saveQueryValue(query_rows_p3, query_col_start_p3, query_data_p3, sheet_data)
    saveQueryValue(query_rows_p4, query_col_start_p4, query_data_p4, sheet_data)
    saveQueryValue(query_rows_p5, query_col_start_p5, query_data_p5, sheet_data)
    saveQueryValue(query_rows_ap5, query_col_start_ap5, query_data_ap5, sheet_data)
    saveQueryValue(query_rows_ndcg, query_col_start_ndcg, query_data_ndcg, sheet_data)


  # 각 쿼리별로 P@1 ~ P@5의 평균을 계산하여 새로운 리스트에 저장
  query_averages_p1 = [[0] * 4 for _ in range(10)]  # P@1 평균값을 저장할 10 x 4 배열
  query_averages_p2 = [[0] * 4 for _ in range(10)]  # P@2 평균값을 저장할 10 x 4 배열
  query_averages_p3 = [[0] * 4 for _ in range(10)]  # P@3 평균값을 저장할 10 x 4 배열
  query_averages_p4 = [[0] * 4 for _ in range(10)]  # P@4 평균값을 저장할 10 x 4 배열
  query_averages_p5 = [[0] * 4 for _ in range(10)]  # P@5 평균값을 저장할 10 x 4 배열
  query_averages_ap5 = [[0] * 4 for _ in range(10)]  # aP@5 평균값을 저장할 10 x 4 배열
  query_averages_ndcg = [[0] * 4 for _ in range(10)]  # ndcg 평균값을 저장할 10 x 4 배열

  for i in range(10):  # 각 쿼리별
    for j in range(4):  # 각 항목별 (KNU, ResNetTrans, BM25, BM25+ResNetTrans)
      query_averages_p1[i][j] = sum(query_data_p1[i][j]) / len(query_data_p1[i][j])  # P@1 평균 계산
      query_averages_p2[i][j] = sum(query_data_p2[i][j]) / len(query_data_p2[i][j])  # P@2 평균 계산
      query_averages_p3[i][j] = sum(query_data_p3[i][j]) / len(query_data_p3[i][j])  # P@3 평균 계산
      query_averages_p4[i][j] = sum(query_data_p4[i][j]) / len(query_data_p4[i][j])  # P@4 평균 계산
      query_averages_p5[i][j] = sum(query_data_p5[i][j]) / len(query_data_p5[i][j])  # P@5 평균 계산
      query_averages_ap5[i][j] = sum(query_data_ap5[i][j]) / len(query_data_ap5[i][j])  # aP@5 평균 계산
      query_averages_ndcg[i][j] = sum(query_data_ndcg[i][j]) / len(query_data_ndcg[i][j])  # ndcg 평균 계산

  print("kkk",query_data_ap5[0][3])
  # 결과 출력
  columns = ['KNU', 'ResNetTrans', 'BM25', 'BM25+ResNetTrans']

  printPK(1, columns, query_averages_p1)
  printPK(2, columns, query_averages_p2)
  printPK(3, columns, query_averages_p3)
  printPK(4, columns, query_averages_p4)
  printPK(5, columns, query_averages_p5)
  printPK("ap", columns, query_averages_ap5)
  printPK("ndcg", columns, query_averages_ndcg)


  #저장
  # 파일 경로지정 (엑셀로 저장할 파일)
  output_excel_file = 'averages.xlsx'

  # 각 항목별로 평균값을 저장할 DataFrame 생성
  columns = ['KNU', 'ResNetTrans', 'BM25', 'BM25+ResNetTrans']
  queries = [f'Query {i+1}' for i in range(10)]

  # P@1 ~ P@5의 평균값을 데이터프레임으로 변환
  df_p1 = pd.DataFrame(query_averages_p1, index=queries, columns=columns)
  df_p2 = pd.DataFrame(query_averages_p2, index=queries, columns=columns)
  df_p3 = pd.DataFrame(query_averages_p3, index=queries, columns=columns)
  df_p4 = pd.DataFrame(query_averages_p4, index=queries, columns=columns)
  df_p5 = pd.DataFrame(query_averages_p5, index=queries, columns=columns)
  df_ap5 = pd.DataFrame(query_averages_ap5, index=queries, columns=columns)
  df_ndcg = pd.DataFrame(query_averages_ndcg, index=queries, columns=columns)


  # ExcelWriter를 사용해 엑셀 파일로 각 P@ 값 저장
  with pd.ExcelWriter(output_excel_file, engine='openpyxl') as writer:
    df_p1.to_excel(writer, sheet_name='P@1 Averages')
    df_p2.to_excel(writer, sheet_name='P@2 Averages')
    df_p3.to_excel(writer, sheet_name='P@3 Averages')
    df_p4.to_excel(writer, sheet_name='P@4 Averages')
    df_p5.to_excel(writer, sheet_name='P@5 Averages')
    df_ap5.to_excel(writer, sheet_name='aP@5 Averages')
    df_ndcg.to_excel(writer, sheet_name='ndcg Averages')

  # 엑셀 파일에 표 스타일 추가
  wb = load_workbook(output_excel_file)

  # 시트별로 표를 추가
  for sheet_name in ['P@1Averages', 'P@2Averages', 'P@3Averages', 'P@4Averages', 'P@5Averages', 'aP@5Averages', 'ndcgAverages']:
    sheet = wb[sheet_name]

    # DataFrame의 크기를 가져와서 표 범위 설정
    rows = sheet.max_row
    cols = sheet.max_column
    ref = f"A1:{chr(64 + cols)}{rows}"

    # 테이블 이름에서 공백을 밑줄로 변경 (또는 다른 방식으로 이름 지정)
    table_name = f"Table_{sheet_name.replace(' ', '_')}"  # 공백 대신 밑줄 사용

    # 표 스타일 설정
    table = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True
    )
    table.tableStyleInfo = style

    # 표 추가
    sheet.add_table(table)

  # 수정된 엑셀 파일 저장
  wb.save(output_excel_file)

  print(f"Excel file saved as {output_excel_file}")
solution()